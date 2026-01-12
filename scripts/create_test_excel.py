"""
Create sample Excel file to test hybrid table format extraction.
"""
import openpyxl
from pathlib import Path

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Products"

# Add headers
ws['A1'] = 'Product'
ws['B1'] = 'Quantity'
ws['C1'] = 'Price'
ws['D1'] = 'Total'

# Add sample data
products = [
    ('Widget A', 100, 10.50, 1050.00),
    ('Widget B', 200, 15.00, 3000.00),
    ('Widget C', 150, 12.00, 1800.00),
    ('Gadget X', 75, 25.00, 1875.00),
    ('Gadget Y', 300, 8.50, 2550.00),
]

for row_idx, (product, qty, price, total) in enumerate(products, start=2):
    ws[f'A{row_idx}'] = product
    ws[f'B{row_idx}'] = qty
    ws[f'C{row_idx}'] = price
    ws[f'D{row_idx}'] = total

# Create another sheet
ws2 = wb.create_sheet("Employees")
ws2['A1'] = 'Name'
ws2['B1'] = 'Age'
ws2['C1'] = 'Department'
ws2['D1'] = 'Salary'

employees = [
    ('John Doe', 30, 'Engineering', 75000),
    ('Jane Smith', 28, 'Marketing', 68000),
    ('Bob Johnson', 35, 'Sales', 72000),
    ('Alice Williams', 32, 'Engineering', 80000),
]

for row_idx, (name, age, dept, salary) in enumerate(employees, start=2):
    ws2[f'A{row_idx}'] = name
    ws2[f'B{row_idx}'] = age
    ws2[f'C{row_idx}'] = dept
    ws2[f'D{row_idx}'] = salary

# Save file
output_path = Path('backend/data/documents/test_sample.xlsx')
output_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(output_path)

print(f"✓ Created test Excel file: {output_path}")
print(f"  - Sheet 1: Products ({len(products)} rows)")
print(f"  - Sheet 2: Employees ({len(employees)} rows)")
